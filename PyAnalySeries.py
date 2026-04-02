#!/usr/bin/env python

#========================================================================================
# Author: Patrick Brockmann CEA/DRF/LSCE - November 2024
#========================================================================================

import sys
import os
import datetime
import re
import copy
from pathlib import Path
import platform

import numpy as np
import pandas as pd

from resources.qt_compat import *

from resources.preferencesDialog import preferencesDialog 

from resources.misc import *
from resources.CustomQColorDialog import CustomQColorDialog 

from resources.displaySingleSeriesWindow import displaySingleSeriesWindow
from resources.displayTogetherSeriesWindow import displayTogetherSeriesWindow
from resources.displayStackedSeriesWindow import displayStackedSeriesWindow

from resources.defineFilterWindow import defineFilterWindow
from resources.displayFilterWindow import displayFilterWindow

from resources.defineInterpolationWindow import defineInterpolationWindow
from resources.displayInterpolationWindow import displayInterpolationWindow

from resources.defineSamplingWindow import defineSamplingWindow
from resources.displaySamplingWindow import displaySamplingWindow

from resources.CustomQTableWidget import CustomQTableWidget

from resources.importDataWindow import importDataWindow

from resources.defineRandomSeriesWindow import defineRandomSeriesWindow
from resources.defineInsolationAstroSeriesWindow import defineInsolationAstroSeriesWindow
from resources.defineSinusoidalSeriesWindow import defineSinusoidalSeriesWindow

from resources.computeAggregateWindow import computeAggregateWindow
from resources.computeDetrendWindow import computeDetrendWindow
from resources.computePowerSpectrumWindow import computePowerSpectrumWindow
from resources.computeFrequencyFilterWindow import computeFrequencyFilterWindow

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

#========================================================================================
if len(sys.argv[1:]) >= 1:
    filesName = sys.argv[1:]
else:
    filesName = None

#========================================================================================
version = 'v6.27'

open_ws = {}
open_displayWindows = {} 
open_filterWindows = {} 
open_samplingWindows = {} 
open_interpolationWindows = {} 

open_importWindow = {}
open_randomSeriesWindow = {}
open_insolationAstroSeriesWindow = {}
open_sinusoidalSeriesWindow = {}

open_aggregateWindows = {}
open_detrendWindows = {}
open_powerSpectrumWindows = {}
open_frequencyFilterWindows = {}

#========================================================================================
def colorize_item(item, color_name, alpha=100):

    tree_widget.blockSignals(True)

    if not color_name:
        brush = QBrush()
    else:
        row = tree_widget.indexOfTopLevelItem(item)
        if row == -1 and item.parent() is not None:
            row = item.parent().indexOfChild(item)

        base_color = QColor("#ffffff") if row % 2 == 0 else QColor("#f0f0f0")
        overlay = QColor(color_name)
        overlay.setAlpha(alpha)
        r = (overlay.red()   * alpha + base_color.red()   * (255 - alpha)) // 255
        g = (overlay.green() * alpha + base_color.green() * (255 - alpha)) // 255
        b = (overlay.blue()  * alpha + base_color.blue()  * (255 - alpha)) // 255
        final_color = QColor(r, g, b)
        brush = QBrush(final_color)

    for col in range(tree_widget.columnCount()):
        item.setBackground(col, brush)
    
    tree_widget.blockSignals(False)

#========================================================================================
def populate_tree_widget(fileName, itemDict_list):
    global open_ws

    tree_widget.blockSignals(True)

    ws_icon = QIcon(str(app_dir / 'resources' / 'icon_folder.png'))

    ws_item = QTreeWidgetItem(tree_widget)
    ws_item.setIcon(0, ws_icon)
    ws_item.setText(0, fileName)
    ws_item.setToolTip(0, fileName)
    ws_item.setExpanded(True)
    ws_item.setFlags(ws_item.flags() & ~ItemIsSelectable)
    open_ws[id(ws_item)] = ws_item.text(0)

    for itemDict in itemDict_list:
        add_item_tree_widget(ws_item, itemDict, mark=False)

    unmark_ws(ws_item)

    tree_widget.blockSignals(False)

    return ws_item

#========================================================================================
def add_item_tree_widget(ws_item, itemDict, position=None, mark=True, update=True):

    tree_widget.blockSignals(True)

    icon_series = QIcon(str(app_dir / 'resources' / 'icon_series.png'))
    icon_seriesReplicates = QIcon(str(app_dir / 'resources' / 'icon_seriesReplicates.png'))

    icon_seriesPSD = QIcon(str(app_dir / 'resources' / 'icon_seriesPSD.png'))
    icon_apply = QIcon(str(app_dir / 'resources' / 'icon_apply.png'))

    item = QTreeWidgetItem()
    item.setFlags(item.flags() & ~ItemIsDropEnabled)

    if itemDict['Type'].startswith('Series'):            # to be backward compatible with Serie and now Series
        Series = itemDict['Series']
        if itemDict['Type'] == 'Series PSD':
            item.setIcon(0, icon_seriesPSD)
        elif Series.index.duplicated().any():
            item.setIcon(0, icon_seriesReplicates)
        else:
            item.setIcon(0, icon_series)
    elif itemDict['Type'] == 'FILTER':
            item.setIcon(0, icon_apply)
    elif itemDict['Type'] in ['SAMPLE', 'SAMPLING']:
            item.setIcon(0, icon_apply)
    elif itemDict['Type'] == 'INTERPOLATION':
            item.setIcon(0, icon_apply)
    else:
        #print("Error: Type unknown")
        return

    # add item at current_index
    if not ws_item:
        current_index = tree_widget.currentItem()
        if not current_index.parent():
            ws_item = current_index
        else:
            ws_item = current_index.parent()

    if position != None:
        ws_item.insertChild(position, item)
    else:
        ws_item.addChild(item)

    if mark:
        mark_ws(ws_item)            # Mark as to be saved

    tree_widget.blockSignals(True)

    item.setData(0, Qt.ItemDataRole.UserRole, itemDict)

    item.setText(0, itemDict['Name'])
    item.setToolTip(0, itemDict['Name'])
    item.setText(1, itemDict['Type'])
    item.setText(2, itemDict['Id'])

    font = QFont('Courier New')
    item.setFont(2, font)             # format Id

    if itemDict['Type'] == 'INTERPOLATION':
        item.setText(3, itemDict['X1Name'])
        item.setFont(3, font)

    if not itemDict['Type'].startswith('Series'):
        if update: update_items_from_data(item)
        return

    item.setText(3, itemDict['X'])
    item.setText(4, itemDict['Y'])
    item.setFont(3, font)
    item.setFont(4, font)

    buttonColor = QPushButton()
    buttonColor.setFixedWidth(40)
    buttonColor.setStyleSheet(f"background-color: {itemDict['Color']};")
    buttonColor.clicked.connect(lambda: selectColor(buttonColor, item))
    tree_widget.setItemWidget(item, 5, buttonColor)
    item.setTextAlignment(5, AlignLeft | AlignVCenter)

    if update: update_items_from_data(item)

    tree_widget.blockSignals(False)

#========================================================================================
def on_item_changed(item, column):
    global open_ws

    #----------------------------------
    if not item.parent():

        #--------
        if column !=0: return

        #--------
        new_wsName = item.text(0).replace(' *', '')
        old_wsName = open_ws[id(item)]
        #print(f"Want to change name ws: {old_wsName} --> {new_wsName}")

        #--------
        if new_wsName == old_wsName: 
            remark_ws(item)
            return

        #--------
        if new_wsName in open_ws.values():
            QMessageBox.warning(main_window, "Duplicate WS name", f"The ws '{new_wsName}' is already in use. Please choose a unique name.")
            item.setText(0, old_wsName)
            return

        #--------
        if os.path.exists(old_wsName) and is_open(old_wsName):
            QMessageBox.warning(main_window, "WS open in another application", f"The ws '{new_wsName}' is already in use. Please close the file.")
            item.setText(0, old_wsName)
            return

        #--------
        file_path = os.path.dirname(new_wsName)
        if file_path and not os.path.exists(file_path):
            QMessageBox.warning(main_window, "Missing directory", f"Create first the target directory on the file system.")
            item.setText(0, old_wsName)
            return

        #--------
        if os.path.exists(old_wsName):
            os.rename(old_wsName, new_wsName)
        remark_ws(item)

        open_ws[id(item)] = new_wsName

    #----------------------------------
    else: 

        tree_widget.blockSignals(True)

        itemDict = item.data(0, Qt.ItemDataRole.UserRole)
        itemDict['Name'] = item.text(0)
        if 'X' in itemDict.keys(): itemDict['X'] = item.text(3)
        elif 'X1Name' in itemDict.keys(): itemDict['X1Name'] = item.text(3)
        if 'Y' in itemDict.keys(): itemDict['Y'] = item.text(4)
        item.setData(0, Qt.ItemDataRole.UserRole, itemDict)

        mark_ws(item.parent())
        update_items_from_data(item)

        tree_widget.blockSignals(False)

#========================================================================================
def selectColor(buttonColor, series_item):
    seriesDict = series_item.data(0, Qt.ItemDataRole.UserRole)
    starting_color = seriesDict['Color']
    color = CustomQColorDialog.getColor(starting_color)
    if color:
        seriesDict = seriesDict | {'Color': color.name()}
        series_item.setData(0, Qt.ItemDataRole.UserRole, seriesDict)
        buttonColor = tree_widget.itemWidget(series_item, 5)
        buttonColor.setStyleSheet(f"background-color: {seriesDict['Color']}; border: none; border-radius: 3px;")
        update_items_from_data(series_item)
        mark_ws(series_item.parent())

#========================================================================================
def update_items_from_data(ref_item):

    tree_widget.blockSignals(True)

    ref_itemDict = ref_item.data(0, Qt.ItemDataRole.UserRole)

    allItems = tree_widget.get_children()

    #print('---', ref_itemDict['Id'], ref_item.parent().text(0))
    
    for item in allItems:

        itemDict = item.data(0, Qt.ItemDataRole.UserRole)

        if itemDict['Id'] == ref_itemDict['Id']:
            #print('----------update', itemDict['Id'], item.parent().text(0))

            before = copy.deepcopy(itemDict)

            item.setText(0, ref_itemDict['Name'])
            if 'X' in ref_itemDict.keys(): item.setText(3, ref_itemDict['X'])
            if 'X1Name' in ref_itemDict.keys(): item.setText(3, ref_itemDict['X1Name'])
            if 'Y' in ref_itemDict.keys(): item.setText(4, ref_itemDict['Y'])
            item.setData(0, Qt.ItemDataRole.UserRole, ref_itemDict)

            if  itemDict['Type'].startswith('Series'):
                buttonColor = tree_widget.itemWidget(item, 5)
                buttonColor.setStyleSheet(f"background-color: {ref_itemDict['Color']}; border: none; border-radius: 3px;")
      
            # Mark if item has changed
            itemDict = item.data(0, Qt.ItemDataRole.UserRole)
            if (before | {'Series': 0}) != (itemDict | {'Series': 0}):
                mark_ws(item.parent())

            tree_widget.blockSignals(True)
                
    tree_widget.blockSignals(False)

#========================================================================================
def load_WorkSheet(fileName):

    if not os.path.exists(fileName): return

    if fileName in open_ws.values():
        msg = f'{fileName} already loaded'
        main_window.statusBar().showMessage(msg, 5000)
        return 

    main_window.statusBar().showMessage(fileName + ' loading', 5000)
    QApplication.processEvents()

    #--------------------------------------------------------------------
    sheetNames = pd.read_excel(fileName, sheet_name=None).keys()

    itemDict_list = []

    for sheetName in sheetNames:

        #-------------------------------------
        if sheetName.startswith(('Serie Id-', 'Series Id-')):

            try:
                df = pd.read_excel(fileName, sheet_name=sheetName, na_filter=False)

                color = QColor(df['Color'][0])
                if color.isValid():
                    Color = df['Color'][0]
                else:
                    Color = generate_color()

                history = df['History'][0]
                history = re.sub(r'^(<br\s*/?>)', '', history, flags=re.IGNORECASE)     # to correct old comments
                history = re.sub(r'\bserie\b', 'series', history)                       # to correct serie to series
                history = re.sub(r'\bSerie\b', 'Series', history)                       # to correct Serie to Series

                type = df['Type'][0] 
                type = re.sub(r'\bSerie\b', 'Series',  type)                            # to correct Serie to Series

                n = effective_length_from_index(df.iloc[:, 0])
                x = addNanList(df.iloc[:, 0], length=n)
                y = addNanList(df.iloc[:, 1], length=n)

                aDict = {
                    'Id': 'Id-' + sheetName.split('Id-')[1],
                    'Type': type,
                    'Name': df['Name'][0],
                    'X':  df.columns[0],
                    'Y':  df.columns[1],
                    'Color': Color,
                    'Date': df['Date'][0] if 'Date' in df.columns else '',
                    'Comment': df['Comment'][0],
                    'History': history,
                    'Series': pd.Series(y, index=x)
                }

                if 'InterpolationMode' in df.columns:
                    aDict = aDict | {
                        'InterpolationMode': df['InterpolationMode'][0],
                        'X1Coords': cleanSpaceList(df['X1Coords']),
                        'X2Coords': cleanSpaceList(df['X2Coords']),
                        'XOriginal': df.columns[11],
                        'XOriginalValues': addNanList(df.iloc[:,11])
                    }

                itemDict_list.append(aDict)

            except:
                msg = f"The file '{fileName}' contains a series that is wrongly formatted in {sheetName} sheet."
                QMessageBox.critical(main_window, "Load file", msg)
                main_window.statusBar().showMessage(msg, 5000)
                QApplication.processEvents()

        #-------------------------------------
        elif sheetName.startswith(('FILTER Id-', 'SAMPLE Id-', 'SAMPLING Id-')):
            
            try:
                df = pd.read_excel(fileName, sheet_name=sheetName, na_filter=False)

                history = df['History'][0]
                history = re.sub(r'^(<br\s*/?>)', '', history, flags=re.IGNORECASE)     # to correct old comments
                history = re.sub(r'\bserie\b', 'series', history)                       # to correct serie to series
                history = re.sub(r'\bSerie\b', 'Series', history)                       # to correct Serie to Series
                history = re.sub(r'\bSAMPLE\b', 'SAMPLING', history)                    # to correct SAMPLE to SAMPLING

                type = df['Type'][0] 
                type = re.sub(r'\bSAMPLE\b', 'SAMPLING',  type)                            # to correct SAMPLE to SAMPLING

                aDict = {
                        'Id': 'Id-' + sheetName.split('Id-')[1],
                        'Type': type,
                        'Name': df['Name'][0],
                        'Parameters': str(df['Parameters'][0]),
                        'Date': df['Date'][0] if 'Date' in df.columns else '',
                        'Comment': df['Comment'][0],
                        'History': history
                }

                if 'XCoords' in df.columns:                 # for SAMPLING
                    aDict = aDict | {
                        'XCoords': addNanList(df['XCoords'])
                    }

                itemDict_list.append(aDict)

            except:
                msg = f"The file '{fileName}' contains a FILTER/SAMPLING that is wrongly formatted in {sheetName} sheet."
                QMessageBox.critical(main_window, "Load file", msg)
                main_window.statusBar().showMessage(msg, 5000)
                QApplication.processEvents()

        #-------------------------------------
        elif sheetName.startswith('INTERPOLATION Id-'):

            try:
                df = pd.read_excel(fileName, sheet_name=sheetName, na_filter=False)

                history = df['History'][0]
                history = re.sub(r'^(<br\s*/?>)', '', history, flags=re.IGNORECASE)     # to correct old comments
                history = re.sub(r'\bserie\b', 'series', history)                       # to correct serie to series
                history = re.sub(r'\bSerie\b', 'Series', history)                       # to correct Serie to Series

                aDict = {
                        'Id': 'Id-' + sheetName.split('INTERPOLATION Id-')[1],
                        'Type': 'INTERPOLATION',
                        'Name': df['Name'][0],
                        'X1Coords': cleanSpaceList(df['X1Coords']),
                        'X2Coords': cleanSpaceList(df['X2Coords']),
                        'X1Name': df['X1Name'][0],
                        'Date': df['Date'][0] if 'Date' in df.columns else '',
                        'Comment': df['Comment'][0],
                        'History': history
                }

                itemDict_list.append(aDict)

            except:
                msg = f"The file '{fileName}' contains an INTERPOLATION that is wrongly formatted in {sheetName} sheet."
                QMessageBox.critical(main_window, "Load file", msg)
                main_window.statusBar().showMessage(msg, 5000)
                QApplication.processEvents()


    #--------------------------------------------------------------------
    if len(itemDict_list) == 0:
        msg = f"The file '{fileName}' has not been recognized as a valid PyAnalySeries worksheet."
        QMessageBox.critical(main_window, "Load file", msg)
        return

    #--------------------------------------------------------------------
    base_dir = os.getcwd()
    fileName = os.path.relpath(fileName, base_dir)          # get relative path
    populate_tree_widget(fileName, itemDict_list)

    #--------------------------------------------------------------------
    settings = QSettings("MyPythonApps", "PyAnalySeries")
    last_used_dir = QFileInfo(fileName).absolutePath()
    settings.setValue("lastDir", last_used_dir)

    #--------------------------------------------------------------------
    main_window.statusBar().showMessage(fileName + ' loaded', 5000)

#========================================================================================
def new_WorkSheet():

    settings = QSettings("MyPythonApps", "PyAnalySeries")
    last_dir = settings.value("lastDir", os.getcwd(), type=str)

    base_dir = os.getcwd()
    rel_dir = os.path.relpath(last_dir, base_dir)          # get relative path

    fileNameTemplate = os.path.join(rel_dir, 'new_{:02d}.xlsx')
    counterFilename = 1
    while (os.path.exists(fileNameTemplate.format(counterFilename))) or \
          (fileNameTemplate.format(counterFilename) in open_ws.values()):
        counterFilename += 1
    fileName = fileNameTemplate.format(counterFilename)

    ws_item = populate_tree_widget(fileName, [])
    mark_ws(ws_item)
    tree_widget.setCurrentItem(ws_item)
    tree_widget.clearSelection()

#========================================================================================
def open_WorkSheet():

    settings = QSettings("MyPythonApps", "PyAnalySeries")
    last_dir = settings.value("lastDir", os.getcwd(), type=str)
    #last_dir = settings.value("lastDir", "", type=str)

    filesName, _ = QFileDialog.getOpenFileNames(
        main_window, "Open Excel File", last_dir, "Excel Files (*.xlsx)"
    )

    if filesName:
        last_used_dir = QFileInfo(filesName[0]).absolutePath()
        settings.setValue("lastDir", last_used_dir)

        for fileName in filesName: 
            print('Loading...', fileName)
            load_WorkSheet(fileName)

#========================================================================================
def autofit_columns(worksheet):

    for col in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[column_letter].width = max_length + 5 

#========================================================================================
def save_WorkSheet(ws_item):

    outFile = ws_item.text(0).replace(" *", "")

    #-----------------------
    if os.path.exists(outFile):
        wb = load_workbook(outFile)
    else:
        wb = Workbook()

    #----------------------------------
    for sheetName in wb.sheetnames:
        # remove sheetname 'Serie Idxxxxxxxx' (wrong spell)
        if sheetName.startswith(('Serie Id', 'Series Id', 'SAMPLE Id', 'SAMPLING Id', 'FILTER Id', 'INTERPOLATION Id')) or \
           sheetName == 'Information' or \
           sheetName == 'Sheet':
            del wb[sheetName]

    #----------------------------------
    try:
        sheetName = f'Information'
        ws = wb.create_sheet(title=sheetName)

        ws.cell(row=1, column=1, value=f'Created with PyAnalyseries {version}')

        text = "This file has been created with PyAnalySeries software."
        ws.cell(row=3, column=1, value=text)
        text = "Do not modify or accordingly with documentation."
        ws.cell(row=4, column=1, value=text)
        text = datetime.datetime.now().strftime("Saved %Y/%m/%d at %H:%M:%S")
        ws.cell(row=6, column=1, value=text)

        #----------------------------------
        for n in range(ws_item.childCount()):

            item = ws_item.child(n)
            itemDict = item.data(0, Qt.ItemDataRole.UserRole)

            #-----------------------
            if itemDict["Type"].startswith('Series'):
                
                sheetName = f'Series {itemDict["Id"]}'
                ws = wb.create_sheet(title=sheetName)

                ws.cell(row=1, column=1, value=itemDict['X'])
                ws.cell(row=1, column=2, value=itemDict['Y'])
                ws.cell(row=1, column=3, value='Type')
                ws.cell(row=1, column=4, value='Name')
                ws.cell(row=1, column=5, value='Color')
                ws.cell(row=1, column=6, value='Date')
                ws.cell(row=1, column=7, value='Comment')
                ws.cell(row=1, column=8, value='History')

                for i, (index, value) in enumerate(itemDict['Series'].sort_index().items(), start=2):            # force sort on index
                    ws.cell(row=i, column=1, value=index)
                    ws.cell(row=i, column=2, value=value)
                ws.cell(row=2, column=3, value=itemDict['Type'])
                ws.cell(row=2, column=4, value=itemDict['Name'])
                ws.cell(row=2, column=5, value=itemDict['Color'])
                ws.cell(row=2, column=6, value=itemDict['Date'])
                ws.cell(row=2, column=7, value=itemDict['Comment'])
                ws.cell(row=2, column=8, value=itemDict['History'])
        
                if 'InterpolationMode' in itemDict:
                    ws.cell(row=1, column=9, value='InterpolationMode')
                    ws.cell(row=1, column=10, value='X1Coords')
                    ws.cell(row=1, column=11, value='X2Coords')
                    ws.cell(row=1, column=12, value=itemDict['XOriginal'])

                    ws.cell(row=2, column=9, value=itemDict['InterpolationMode'])
                    for i, value in enumerate(itemDict['X1Coords'], start=2):
                        ws.cell(row=i, column=10, value=value)
                    for i, value in enumerate(itemDict['X2Coords'], start=2):
                        ws.cell(row=i, column=11, value=value)
                    for i, value in enumerate(itemDict['XOriginalValues'], start=2):
                        ws.cell(row=i, column=12, value=value)

            #-----------------------
            elif itemDict["Type"] in ['FILTER', 'SAMPLE', 'SAMPLING']:

                sheetName = f'{itemDict["Type"]} {itemDict["Id"]}'
                ws = wb.create_sheet(title=sheetName)

                ws.cell(row=1, column=1, value='Type')
                ws.cell(row=1, column=2, value='Name')
                ws.cell(row=1, column=3, value='Parameters')
                ws.cell(row=1, column=4, value='Date')
                ws.cell(row=1, column=5, value='Comment')
                ws.cell(row=1, column=6, value='History')
                
                ws.cell(row=2, column=1, value=itemDict['Type'])
                ws.cell(row=2, column=2, value=itemDict['Name'])
                ws.cell(row=2, column=3, value=itemDict['Parameters'])
                ws.cell(row=2, column=4, value=itemDict['Date'])
                ws.cell(row=2, column=5, value=itemDict['Comment'])
                ws.cell(row=2, column=6, value=itemDict['History'])
                
                if 'XCoords' in itemDict:
                    ws.cell(row=1, column=7, value='XCoords')
                    for i, value in enumerate(itemDict['XCoords'], start=2):
                        ws.cell(row=i, column=7, value=value)

            #-----------------------
            elif itemDict["Type"] == 'INTERPOLATION':

                sheetName = f'{itemDict["Type"]} {itemDict["Id"]}'
                ws = wb.create_sheet(title=sheetName)

                ws.cell(row=1, column=1, value='X1Coords')
                ws.cell(row=1, column=2, value='X2Coords')
                ws.cell(row=1, column=3, value='X1Name')
                ws.cell(row=1, column=4, value='Type')
                ws.cell(row=1, column=5, value='Name')
                ws.cell(row=1, column=6, value='Date')
                ws.cell(row=1, column=7, value='Comment')
                ws.cell(row=1, column=8, value='History')

                for i, value in enumerate(itemDict['X1Coords'], start=2):
                    ws.cell(row=i, column=1, value=value)
                for i, value in enumerate(itemDict['X2Coords'], start=2):
                    ws.cell(row=i, column=2, value=value)
                ws.cell(row=2, column=3, value=itemDict['X1Name'])
                ws.cell(row=2, column=4, value=itemDict['Type'])
                ws.cell(row=2, column=5, value=itemDict['Name'])
                ws.cell(row=2, column=6, value=itemDict['Date'])
                ws.cell(row=2, column=7, value=itemDict['Comment'])
                ws.cell(row=2, column=8, value=itemDict['History'])
                
            #-----------------------

        for sheet in wb.worksheets:
            autofit_columns(sheet)

        wb.save(outFile)
        return True 

    #-----------------------
    except:
        return False 

#========================================================================================
def save_WorkSheetCurrent():

    current_item = tree_widget.currentItem()

    if not current_item:             # no ws present
        return

    ws_item = current_item.parent() if current_item.parent() else current_item
    ws_name = ws_item.text(0)

    if ws_item.data(0, Qt.ItemDataRole.UserRole): 
        ws_name = ws_name.replace(" *", "")
        if save_WorkSheet(ws_item):
            unmark_ws(ws_item)
            main_window.statusBar().showMessage(f'Worksheet {ws_name} saved', 3000)
        else:
            main_window.statusBar().showMessage(f'Error when saving {ws_name}', 3000)
    else:
        main_window.statusBar().showMessage(f'Worksheet {ws_name} has no unsaved changes', 3000)

#========================================================================================
def save_WorkSheets():

    display_error = False
    for ws_item in tree_widget.get_parents():
        if ws_item.data(0, Qt.ItemDataRole.UserRole): 
            ws_name = ws_item.text(0).replace(" *", "")
            if save_WorkSheet(ws_item):
                unmark_ws(ws_item)
            else:
                main_window.statusBar().showMessage(f'Error when saving {ws_name}', 3000)
                display_error = True

    if not display_error:
        main_window.statusBar().showMessage('Worksheets saved', 3000)

#========================================================================================
def import_Data():
    global open_importWindow

    current_index = tree_widget.currentItem()
    if not current_index:
        new_WorkSheet()
    
    Id_importWindow = '123456'

    if open_importWindow:
        importWindow = open_importWindow[Id_importWindow]
        importWindow.raise_()
        importWindow.activateWindow()
    else:
        importWindow = importDataWindow(open_importWindow, add_item_tree_widget)
        open_importWindow[Id_importWindow] = importWindow
        importWindow.show()

#========================================================================================
def define_insolationAstroSeries():
    global open_insolationAstroSeriesWindow

    current_index = tree_widget.currentItem()
    if not current_index:
        new_WorkSheet()
    
    Id_insolationAstroSeriesWindow = '123456'

    if open_insolationAstroSeriesWindow:
        insolationAstroSeriesWindow = open_insolationAstroSeriesWindow[Id_insolationAstroSeriesWindow]
        insolationAstroSeriesWindow.raise_()
        insolationAstroSeriesWindow.activateWindow()
    else:
        insolationAstroSeriesWindow = defineInsolationAstroSeriesWindow(open_insolationAstroSeriesWindow, add_item_tree_widget)
        open_insolationAstroSeriesWindow[Id_insolationAstroSeriesWindow] = insolationAstroSeriesWindow
        insolationAstroSeriesWindow.show()

#========================================================================================
def define_sinusoidalSeries():
    global open_sinusoidalSeriesWindow

    current_index = tree_widget.currentItem()
    if not current_index:
        new_WorkSheet()
    
    Id_sinusoidalSeriesWindow = '123456'

    if open_sinusoidalSeriesWindow:
        sinusoidalSeriesWindow = open_sinusoidalSeriesWindow[Id_sinusoidalSeriesWindow]
        sinusoidalSeriesWindow.raise_()
        sinusoidalSeriesWindow.activateWindow()
    else:
        sinusoidalSeriesWindow = defineSinusoidalSeriesWindow(open_sinusoidalSeriesWindow, add_item_tree_widget)
        open_sinusoidalSeriesWindow[Id_sinusoidalSeriesWindow] = sinusoidalSeriesWindow
        sinusoidalSeriesWindow.show()

#========================================================================================
def define_randomSeries():
    global open_randomSeriesWindow

    current_index = tree_widget.currentItem()
    if not current_index:
        new_WorkSheet()
    
    Id_randomSeriesWindow = '123456'

    if open_randomSeriesWindow:
        randomSeriesWindow = open_randomSeriesWindow[Id_randomSeriesWindow]
        randomSeriesWindow.raise_()
        randomSeriesWindow.activateWindow()
    else:
        randomSeriesWindow = defineRandomSeriesWindow(open_randomSeriesWindow, add_item_tree_widget)
        open_randomSeriesWindow[Id_randomSeriesWindow] = randomSeriesWindow
        randomSeriesWindow.show()

#========================================================================================
def compute_aggregate():
    global open_aggregateWindows

    items = get_unique_selected_items(tree_widget)
    items_selected = []                             # select only series
    for item in items:
        seriesDict = item.data(0, Qt.ItemDataRole.UserRole)
        if  seriesDict['Type'].startswith('Series'): 
            items_selected.append(item)

    if len(items_selected) != 1 : 
        main_window.statusBar().showMessage('Please select only 1 series', 5000)
        return

    #-------------------------------------------------------------
    Id_aggregateWindow = generate_Id()

    if Id_aggregateWindow in open_aggregateWindows:
        aggregateWindow = open_aggregateWindows[Id_aggregateWindow]
        aggregateWindow.raise_()
        aggregateWindow.activateWindow()
    else:
        aggregateWindow = computeAggregateWindow(Id_aggregateWindow, open_aggregateWindows, item, add_item_tree_widget)
        open_aggregateWindows[Id_aggregateWindow] = aggregateWindow
        aggregateWindow.show()

    #-------------------------------------------------------------
    main_window.setFocus()                  # replace selection
    tree_widget.clearSelection()
    item.setSelected(True)

#========================================================================================
def compute_detrend():
    global open_detrendWindows

    items = get_unique_selected_items(tree_widget)
    items_selected = []                             # select only series
    for item in items:
        seriesDict = item.data(0, Qt.ItemDataRole.UserRole)
        if  seriesDict['Type'].startswith('Series'): 
            items_selected.append(item)

    if len(items_selected) != 1 : 
        main_window.statusBar().showMessage('Please select only 1 series', 5000)
        return

    #-------------------------------------------------------------
    Id_detrendWindow = generate_Id()

    if Id_detrendWindow in open_detrendWindows:
        detrendWindow = open_detrendWindows[Id_detrendWindow]
        detrendWindow.raise_()
        detrendWindow.activateWindow()
    else:
        detrendWindow = computeDetrendWindow(Id_detrendWindow, open_detrendWindows, item, add_item_tree_widget)
        open_detrendWindows[Id_detrendWindow] = detrendWindow
        detrendWindow.show()

    #-------------------------------------------------------------
    main_window.setFocus()                  # replace selection
    tree_widget.clearSelection()
    item.setSelected(True)

#========================================================================================
def compute_powerSpectrum():
    global open_powerSpectrumWindows

    items = get_unique_selected_items(tree_widget)
    items_selected = []                             # select only series
    for item in items:
        seriesDict = item.data(0, Qt.ItemDataRole.UserRole)
        if  seriesDict['Type'].startswith('Series'): 
            items_selected.append(item)

    if len(items_selected) != 1 : 
        main_window.statusBar().showMessage('Please select only 1 series', 5000)
        return

    #-------------------------------------------------------------
    Id_powerSpectrumWindow = generate_Id()

    if Id_powerSpectrumWindow in open_powerSpectrumWindows:
        powerSpectrumWindow = open_powerSpectrumWindows[Id_powerSpectrumWindow]
        powerSpectrumWindow.raise_()
        powerSpectrumWindow.activateWindow()
    else:
        powerSpectrumWindow = computePowerSpectrumWindow(Id_powerSpectrumWindow, open_powerSpectrumWindows, item, add_item_tree_widget)
        open_powerSpectrumWindows[Id_powerSpectrumWindow] = powerSpectrumWindow
        powerSpectrumWindow.show()

    #-------------------------------------------------------------
    main_window.setFocus()                  # replace selection
    tree_widget.clearSelection()
    item.setSelected(True)

#========================================================================================
def compute_frequencyFilter():
    global open_frequencyFilterWindows

    items = get_unique_selected_items(tree_widget)
    items_selected = []                             # select only series
    for item in items:
        seriesDict = item.data(0, Qt.ItemDataRole.UserRole)
        if  seriesDict['Type'].startswith('Series'): 
            items_selected.append(item)

    if len(items_selected) != 1 : 
        main_window.statusBar().showMessage('Please select only 1 series', 5000)
        return

    #-------------------------------------------------------------
    Id_frequencyFilterWindow = generate_Id()

    if Id_frequencyFilterWindow in open_frequencyFilterWindows:
        frequencyFilterWindow = open_frequencyFilterWindows[Id_frequencyFilterWindow]
        frequencyFilterWindow.raise_()
        frequencyFilterWindow.activateWindow()
    else:
        frequencyFilterWindow = computeFrequencyFilterWindow(Id_frequencyFilterWindow, open_frequencyFilterWindows, item, add_item_tree_widget)
        open_frequencyFilterWindows[Id_frequencyFilterWindow] = frequencyFilterWindow
        frequencyFilterWindow.show()

    #-------------------------------------------------------------
    main_window.setFocus()                  # replace selection
    tree_widget.clearSelection()
    item.setSelected(True)

#========================================================================================
def create_tree_widget():

    size = QApplication.instance().font().pointSize()
    font = QFont('Courier New', size)

    tree_widget = CustomTreeWidget()
    tree_widget.setColumnCount(6)
    tree_widget.setHeaderLabels(["Name", "Type", "Id", "X", "Y", "Color"])
    tree_widget.setColumnWidth(0, 400)
    tree_widget.setColumnWidth(1, 150)
    tree_widget.setColumnWidth(2, 150)
    tree_widget.setColumnWidth(3, 300)
    tree_widget.setColumnWidth(4, 300)
    tree_widget.setColumnWidth(5, 20)

    tree_widget.setAlternatingRowColors(True)
    tree_widget.setTextElideMode(Qt.TextElideMode.ElideRight)
    tree_widget.setSelectionMode(QTreeWidget.SelectionMode.ExtendedSelection)
    tree_widget.setStyleSheet("""
        QTreeView {
            selection-background-color: rgba(0, 120, 215, 80);
            selection-color: black;
        }
    """)
    tree_widget.setFocusPolicy(Qt.FocusPolicy.ClickFocus)
    tree_widget.headerItem().setFont(2, font)
    tree_widget.headerItem().setFont(3, font)
    tree_widget.headerItem().setFont(4, font)

    #---------------------------------------------
    tree_widget.edit_delegate = QStyledItemDelegate(tree_widget)
    
    def create_editor(parent, option, index):
        editor = QLineEdit(parent)
        if index.column() in (3, 4): editor.setFont(font)
        editor.setStyleSheet("color: blue;")
        return editor
    
    tree_widget.edit_delegate.createEditor = create_editor
    tree_widget.setItemDelegateForColumn(0, tree_widget.edit_delegate)
    tree_widget.setItemDelegateForColumn(3, tree_widget.edit_delegate)
    tree_widget.setItemDelegateForColumn(4, tree_widget.edit_delegate)
    
    #---------------------------------------------
    tree_widget.readonly_delegate = QStyledItemDelegate(tree_widget)
    
    def create_readonly_editor(parent, option, index):
        editor = QLineEdit(parent)
        editor.setFont(font)
        editor.setReadOnly(True)
        editor.setStyleSheet("color: red;")
        return editor
    
    tree_widget.readonly_delegate.createEditor = create_readonly_editor
    tree_widget.setItemDelegateForColumn(2, tree_widget.readonly_delegate)

    #---------------------------------------------
    tree_widget.setDragEnabled(True)
    tree_widget.setAcceptDrops(True)
    tree_widget.setDropIndicatorShown(True)
    tree_widget.setDragDropMode(QTreeWidget.DragDropMode.InternalMove)
    tree_widget.invisibleRootItem().setFlags(ItemIsEnabled)      # root non droppable

    return tree_widget

#========================================================================================
def update_tree_visuals(tree_widget):

    fm = tree_widget.fontMetrics()
    row_h = fm.height()

    # --- Boutons (via stylesheet)
    button_h = row_h - 2
    tree_widget.setStyleSheet(f"""
    QTreeWidget QPushButton {{
        min-height: {button_h}px;
        max-height: {button_h}px;
        padding: 0px;
        border-radius: 3px;
    }}
    """)

    # --- Icons
    icon_h = row_h - 2 
    tree_widget.setIconSize(QSize(icon_h, icon_h))

#========================================================================================
class FullRowDelegate(QStyledItemDelegate):
    def paint(self, painter, option, index):
        brush = index.data(Qt.ItemDataRole.BackgroundRole)

        if isinstance(brush, QBrush) and index.column() == 0 and option.widget:
            r = option.rect
            full_row = QRect(0, r.y(), option.widget.width(), r.height())
            painter.fillRect(full_row, brush)

        super().paint(painter, option, index)

#========================================================================================
class CustomTreeWidget(QTreeWidget):

    #-----------------------------------
    def __init__(self):
        super().__init__()
        self.clipboard_items = []
        self.setMouseTracking(True)
        self.viewport().setAttribute(WA_Hover, True)
        self.viewport().installEventFilter(self)

        self.setItemDelegate(FullRowDelegate(self))

        # 🔹 Custom QLabel as tooltip replacement
        self.custom_tooltip_enabled = True

        self.custom_tooltip = QLabel(self)
        self.custom_tooltip.setStyleSheet("""
            background-color: lightyellow;
            border: 1px solid lightgray;
            border-radius: 4px;
            padding: 4px;
            font-family: Courier New;
            ul { margin: 0px; }
        """)
        # font-size: 12px;
        self.custom_tooltip.setWindowFlags(ToolTip)
        self.custom_tooltip.setFixedWidth(500)
        self.custom_tooltip.setWordWrap(True)
        self.custom_tooltip.setAlignment(AlignTop | AlignLeft)  # Align top-left
        self.custom_tooltip.hide()

    #-----------------------------------
    def set_custom_tooltip_enabled(self, enabled: bool):
        self.custom_tooltip_enabled = enabled
        if not enabled:
            self.custom_tooltip.hide()

    #-----------------------------------
    def dropEvent(self, event):

        selected = self.selectedItems()
        if len(selected) > 1:
            main_window.statusBar().showMessage('You can only drop one item at a time.', 5000)
            event.ignore()
            return

        dragged_item = self.currentItem()
        x, y = event_pos(event)
        target_item = self.itemAt(int(x), int(y))

        # drag only in same WS
        if not target_item or dragged_item.parent() != target_item.parent():
            event.ignore()
            main_window.statusBar().showMessage('Items can only be dragged within the same worksheet.', 5000)
            return
        
        tree_widget.blockSignals(True)

        # Find the position where to move the drag item
        position = target_item.parent().indexOfChild(target_item)
        # Retrieve the data from the dragged item
        itemDict = dragged_item.data(0, Qt.ItemDataRole.UserRole)
        dragged_item.parent().removeChild(dragged_item)
        # Use the `add_item_tree_widget` function to add the dragged item to the target parent
        add_item_tree_widget(target_item.parent(), itemDict, position, mark=False, update=False)

        # Call the default implementation if the drop is valid
        super().dropEvent(event)

        mark_ws(target_item.parent())

        tree_widget.blockSignals(False)
       
    #-----------------------------------
    def eventFilter(self, obj, event):

        if not self.custom_tooltip_enabled:
            return super().eventFilter(obj, event)

        """ Event filter to detect mouse hover and display a tooltip """
        if event.type() == HoverMove:
            x, y = event_pos(event)
            pos = QPoint(int(x), int(y))
            item = self.itemAt(pos)
            col = self.columnAt(pos.x())

            if item and col == 2:  # Tooltip only for column 1
                data = item.data(0, Qt.ItemDataRole.UserRole)
                if isinstance(data, dict):
                    tooltip_text = '''<style> 
                                             p,ol { margin: 0px 0px 0px 0px; }
                                             .p1 { margin: 0px 0px 0px 20px; }
                                      </style>''';
                    tooltip_text += f'''<p><u>Date :</u><p class='p1'>{data['Date']}
                                        <p><u>History :</u><ol><li>{data['History']}</ol>
                                        <p><u>Comment :</u><p class='p1'>{data['Comment']}'''

                    global_pos = self.viewport().mapToGlobal(pos)
                    self.custom_tooltip.setText(tooltip_text)
                    self.custom_tooltip.move(global_pos + QPoint(10, 10))
                    self.custom_tooltip.adjustSize()
                    self.custom_tooltip.show()
                    return True

            else:
                self.custom_tooltip.hide()

        elif event.type() in [Leave, FocusOut]:
            self.custom_tooltip.hide()

        return super().eventFilter(obj, event)

    #-----------------------------------
    def get_parents(self):
        parents = []
        for i in range(self.topLevelItemCount()):
            parents.append(self.topLevelItem(i))
        return parents

    #-----------------------------------
    def get_children(self):
        children = []
        for i in range(self.topLevelItemCount()):
            parent_item = self.topLevelItem(i)
            for j in range(parent_item.childCount()):
                children.append(parent_item.child(j))
        return children

#========================================================================================
def get_unique_selected_items(tree_widget):
    selected_items = tree_widget.selectedItems()
    unique_ids = set()
    unique_items = []

    for item in selected_items:
        itemDict = item.data(0, Qt.ItemDataRole.UserRole)

        if itemDict['Id'] not in unique_ids:
            unique_ids.add(itemDict['Id'])  # Ajoute l'ID à l'ensemble
            unique_items.append(item)  # Ajoute l'item à la liste unique

    return unique_items

#========================================================================================
def displaySingleSeries_selected_series():
    global open_displayWindows

    items = get_unique_selected_items(tree_widget)
    if len(items) == 0:
        main_window.statusBar().showMessage('Please select at least 1 series', 5000)
        return

    for item in items:
        itemDict = item.data(0, Qt.ItemDataRole.UserRole)

        Id_displayWindow = itemDict['Id']

        if Id_displayWindow in open_displayWindows:
            displayWindow = open_displayWindows[Id_displayWindow]
            displayWindow.raise_()
            displayWindow.activateWindow()
        else:
            if itemDict['Type'].startswith('Series'): 
                displayWindow = displaySingleSeriesWindow(Id_displayWindow, open_displayWindows, item)
            elif itemDict['Type'] == 'FILTER':
                displayWindow = displayFilterWindow(Id_displayWindow, open_displayWindows, item)
            elif itemDict['Type'] == 'SAMPLING':
                displayWindow = displaySamplingWindow(Id_displayWindow, open_displayWindows, item)
            elif itemDict['Type'] == 'INTERPOLATION':
                displayWindow = displayInterpolationWindow(Id_displayWindow, open_displayWindows, item)
            open_displayWindows[Id_displayWindow] = displayWindow
            displayWindow.show()

    #-------------------------------------------------------------
    main_window.setFocus()                  # replace selection
    tree_widget.clearSelection()
    for item in items:
        item.setSelected(True)

#========================================================================================
def displayMultipleSeries_selected_series(overlaid=True):
    global open_displayWindows

    items = get_unique_selected_items(tree_widget)
    if len(items) == 0:
        main_window.statusBar().showMessage('Please select at least 1 series', 5000)
        return
    elif len(items) == 1:                             # If only 1 item selected
        displaySingleSeries_selected_series()
        return

    if len(items) > 8: 
        main_window.statusBar().showMessage('Please select 8 series maximum', 3000)
        return

    items_selected = []                             # select only series
    seriesDicts = []
    for item in items:
        itemDict = item.data(0, Qt.ItemDataRole.UserRole)
        if  itemDict['Type'].startswith('Series'):
            items_selected.append(item)
            seriesDicts.append(itemDict)

    if len(items_selected) < 2:
        main_window.statusBar().showMessage('Please select at least 2 series', 5000)
        return

    #-------------------------------------------------------------
    Id_displayWindow = tuple(seriesDict['Id'] for seriesDict in seriesDicts)

    if Id_displayWindow in open_displayWindows:
        displayWindow = open_displayWindows[Id_displayWindow]
        displayWindow.raise_()
        displayWindow.activateWindow()
    else:
        if overlaid:
            displayWindow = displayTogetherSeriesWindow(Id_displayWindow, open_displayWindows, items_selected)
        else:
            displayWindow = displayStackedSeriesWindow(Id_displayWindow, open_displayWindows, items_selected)
        open_displayWindows[Id_displayWindow] = displayWindow
        displayWindow.show()

    #-------------------------------------------------------------
    main_window.setFocus()                  # replace selection
    tree_widget.clearSelection()
    for item in items_selected:
        item.setSelected(True)

#========================================================================================
def define_filter():
    global open_filterWindows

    items = get_unique_selected_items(tree_widget)
    items_selected = []                             # select only series
    for item in items:
        seriesDict = item.data(0, Qt.ItemDataRole.UserRole)
        if  seriesDict['Type'].startswith('Series'): 
            items_selected.append(item)

    if len(items_selected) != 1 : 
        main_window.statusBar().showMessage('Please select only 1 series', 5000)
        return

    #-------------------------------------------------------------
    Id_filterWindow = generate_Id()

    if Id_filterWindow in open_filterWindows:
        filterWindow = open_filterWindows[Id_filterWindow]
        filterWindow.raise_()
        filterWindow.activateWindow()
    else:
        filterWindow = defineFilterWindow(Id_filterWindow, open_filterWindows, item, add_item_tree_widget)
        open_filterWindows[Id_filterWindow] = filterWindow
        filterWindow.show()

    #-------------------------------------------------------------
    main_window.setFocus()                  # replace selection
    tree_widget.clearSelection()
    item.setSelected(True)

#========================================================================================
def apply_filter():

    items = get_unique_selected_items(tree_widget)
    itemSeries_selected = []
    itemFilters_selected = []
    for item in items:
        itemDict = item.data(0, Qt.ItemDataRole.UserRole)
        if  itemDict['Type'].startswith('Series'): 
            itemSeries_selected.append(item)
        elif itemDict['Type'] == 'FILTER':
            itemFilters_selected.append(item)

    if len(itemFilters_selected) != 1 or len(itemSeries_selected) < 1:
        main_window.statusBar().showMessage('Please select 1 FILTER and at least 1 series', 5000)
        return
       
    #-------------------------------------------------------------
    tree_widget.clearSelection()
    itemFilter = itemFilters_selected[0]
    colorize_item(itemFilter, 'red')
    for item in itemSeries_selected:
        colorize_item(item, 'green')

    reply = QMessageBox.question(
        main_window, 
        "Apply filter confirmation",
        "Do you want to apply filter on selected series ?",
        MessageBoxYes | MessageBoxNo,
        MessageBoxNo
    )
    
    if reply == MessageBoxNo:
        for item in items:
            colorize_item(item, None)
        tree_widget.clearSelection()
        return

    #-------------------------------------------------------------
    filterDict = itemFilter.data(0, Qt.ItemDataRole.UserRole)
    filter_window_size = int(filterDict['Parameters'])

    for item in itemSeries_selected:
        seriesDict = item.data(0, Qt.ItemDataRole.UserRole)
        series = seriesDict['Series']
        series = series.groupby(series.index).mean()

        # to keep replicates
        series_XMean = defineFilterWindow.moving_average(series, window_size=filter_window_size)
        series = seriesDict['Series']
        series_filtered = series_XMean.reindex(series.index)

        filtered_Id = generate_Id()
        filtered_seriesDict = seriesDict | {'Id': filtered_Id,
            'Type': 'Series filtered',
            'Series': series_filtered,
            'Color': generate_color(exclude_color=seriesDict['Color']),
            'Date': datetime.datetime.now().strftime("Created %Y/%m/%d at %H:%M:%S"),
            'History': append_to_htmlText(seriesDict['History'], 
                f'Series <i><b>{seriesDict["Id"]}</i></b> filtered with FILTER <i><b>{filterDict["Id"]}</i></b> with a moving average of size {filter_window_size}<BR>---> series <i><b>{filtered_Id}</b></i>'),
            'Comment': ''
        }
        ws_item = item.parent()
        position = ws_item.indexOfChild(item)
        add_item_tree_widget(ws_item, filtered_seriesDict, position+1)

    #-------------------------------------------------------------
    main_window.setFocus()                  # replace selection
    tree_widget.clearSelection()
    for item in itemSeries_selected + itemFilters_selected:
        colorize_item(item, None)
        item.setSelected(True)

#========================================================================================
def define_sampling():
    global open_samplingWindows

    items = get_unique_selected_items(tree_widget)
    items_selected = []                             # select only series
    for item in items:
        seriesDict = item.data(0, Qt.ItemDataRole.UserRole)
        if  seriesDict['Type'].startswith('Series'): 
            items_selected.append(item)

    if len(items_selected) == 0 or len(items_selected) > 2 : 
        main_window.statusBar().showMessage('Please select at least 1 series (2nd possible for sampling reference)', 5000)
        return

    #-------------------------------------------------------------
    Id_samplingWindow = generate_Id()

    if Id_samplingWindow in open_samplingWindows:
        samplingWindow = open_samplingWindows[Id_samplingWindow]
        samplingWindow.raise_()
        samplingWindow.activateWindow()
    else:
        samplingWindow = defineSamplingWindow(Id_samplingWindow, open_samplingWindows, items_selected, add_item_tree_widget)
        open_samplingWindows[Id_samplingWindow] = samplingWindow
        samplingWindow.show()

    #-------------------------------------------------------------
    main_window.setFocus()                  # replace selection
    tree_widget.clearSelection()
    item.setSelected(True)

#========================================================================================
def apply_sampling():

    items = get_unique_selected_items(tree_widget)
    itemSeries_selected = []
    itemSamples_selected = []
    for item in items:
        itemDict = item.data(0, Qt.ItemDataRole.UserRole)
        if  itemDict['Type'].startswith('Series'): 
            itemSeries_selected.append(item)
        elif itemDict['Type'] == 'SAMPLING':
            itemSamples_selected.append(item)

    if len(itemSamples_selected) != 1 or len(itemSeries_selected) < 1:
        main_window.statusBar().showMessage('Please select 1 SAMPLING and at least 1 series', 5000)
        return
       
    #-------------------------------------------------------------
    tree_widget.clearSelection()
    itemFilter = itemSamples_selected[0]
    colorize_item(itemFilter, 'red')
    for item in itemSeries_selected:
        colorize_item(item, 'green')

    reply = QMessageBox.question(
        main_window, 
        "Apply sampling confirmation",
        "Do you want to apply sampling on selected series ?",
        MessageBoxYes | MessageBoxNo,
        MessageBoxNo
    )
    
    if reply == MessageBoxNo:
        for item in items:
            colorize_item(item, None)
        tree_widget.clearSelection()
        return

    #-------------------------------------------------------------
    samplingDict = itemFilter.data(0, Qt.ItemDataRole.UserRole)

    for item in itemSeries_selected:
        seriesDict = item.data(0, Qt.ItemDataRole.UserRole)
        series = seriesDict['Series']
        series = series.groupby(series.index).mean()

        try:
            if 'XCoords' in samplingDict.keys():
                param1_str, param2_str = samplingDict['Parameters'].split(';')
                sampling_kind = param1_str.strip()
                sampling_integrated = str_to_bool(param2_str.strip())
                sampling_index =  samplingDict['XCoords']
                if sampling_integrated: 
                    textHistory = f'using x values and {sampling_kind} interpolation with integration'
                else:
                    textHistory = f'using x values and {sampling_kind} interpolation'
            else:
                param1_str, param2_str, param3_str = samplingDict['Parameters'].split(';')
                sampling_step = float(param1_str.strip())
                sampling_kind = param2_str.strip()
                sampling_integrated = str_to_bool(param3_str.strip())
                index_min = series.index.min()
                index_max = series.index.max()
                index_min = np.ceil(index_min / sampling_step) * sampling_step
                index_max = np.floor(index_max / sampling_step) * sampling_step
                sampling_index = np.arange(index_min, index_max + sampling_step, sampling_step)
                if sampling_integrated: 
                    textHistory = f'every {sampling_step} and {sampling_kind} interpolation with integration'
                else:
                    textHistory = f'every {sampling_step} and {sampling_kind} interpolation'

            sampled_Id = generate_Id()
            sampled_seriesDict = seriesDict | {'Id': sampled_Id,
                'Type': 'Series sampled',
                'Series': defineSamplingWindow.sampling(series, sampling_index, kind=sampling_kind, integrated=sampling_integrated),
                'Color': generate_color(exclude_color=seriesDict['Color']),
                'Date': datetime.datetime.now().strftime("Created %Y/%m/%d at %H:%M:%S"),
                'History': append_to_htmlText(seriesDict['History'], 
                    f'Series <i><b>{seriesDict["Id"]}</i></b> sampled {textHistory} with SAMPLING <i><b>{samplingDict["Id"]}</i></b><BR>---> series <i><b>{sampled_Id}</b></i>'),
                'Comment': ''
            }
            ws_item = item.parent()
            position = ws_item.indexOfChild(item)
            add_item_tree_widget(ws_item, sampled_seriesDict, position+1)

        except:
            msg = f"Problem when applying Sampling. Delete the uncorrect SAMPLING and redefine a new one."
            main_window.statusBar().showMessage(msg, 5000)
            QApplication.processEvents()

    #-------------------------------------------------------------
    main_window.setFocus()                  # replace selection
    tree_widget.clearSelection()
    for item in itemSeries_selected + itemSamples_selected:
        colorize_item(item, None)
        item.setSelected(True)

#========================================================================================
def define_interpolation():
    global open_interpolationWindows

    items = get_unique_selected_items(tree_widget)
    itemSeries_selected = []
    itemInterpolations_selected = []
    for item in items:
        itemDict = item.data(0, Qt.ItemDataRole.UserRole)
        if  itemDict['Type'].startswith('Series'): 
            itemSeries_selected.append(item)
        elif itemDict['Type'] == 'INTERPOLATION':
            itemInterpolations_selected.append(item)

    if not (len(itemInterpolations_selected) <= 1 and len(itemSeries_selected) >= 2):
        main_window.statusBar().showMessage('Please select at least 2 series and optionnaly 1 INTERPOLATION', 5000)
        return
       
    #-------------------------------------------------------------
    if len(itemInterpolations_selected) == 1:
        itemInterpolation = itemInterpolations_selected[0]
        itemDict = itemInterpolation.data(0, Qt.ItemDataRole.UserRole)
        Id_interpolationWindow = itemDict['Id']
    else:
        itemInterpolation = None
        Id_interpolationWindow = generate_Id()
    items = itemSeries_selected

    if Id_interpolationWindow in open_interpolationWindows:
        interpolationWindow = open_interpolationWindows[Id_interpolationWindow]
        interpolationWindow.raise_()
        interpolationWindow.activateWindow()
    else:
        interpolationWindow = defineInterpolationWindow(Id_interpolationWindow, open_interpolationWindows, 
                                itemInterpolation, items, add_item_tree_widget)
        open_interpolationWindows[Id_interpolationWindow] = interpolationWindow
        interpolationWindow.show()

    #-------------------------------------------------------------
    main_window.setFocus()                  # replace selection
    tree_widget.clearSelection()
    for item in itemSeries_selected + itemInterpolations_selected:
        colorize_item(item, None)
        item.setSelected(True)

#========================================================================================
def apply_interpolation(interpolationMode):

    items = get_unique_selected_items(tree_widget)
    itemSeries_selected = []
    itemInterpolations_selected = []
    for item in items:
        itemDict = item.data(0, Qt.ItemDataRole.UserRole)
        if  itemDict['Type'].startswith('Series'): 
            itemSeries_selected.append(item)
        elif itemDict['Type'] == 'INTERPOLATION':
            itemInterpolations_selected.append(item)

    if len(itemInterpolations_selected) != 1 or len(itemSeries_selected) < 1:
        main_window.statusBar().showMessage('Please select 1 INTERPOLATION and at least 1 series', 5000)
        return
       
    #-------------------------------------------------------------
    tree_widget.clearSelection()
    itemInterpolation = itemInterpolations_selected[0]
    colorize_item(itemInterpolation, 'red')
    for item in itemSeries_selected:
        colorize_item(item, 'green')

    reply = QMessageBox.question(
        main_window, 
        "Apply interpolation confirmation",
        "Do you want to apply interpolation on selected series ?",
        MessageBoxYes | MessageBoxNo,
        MessageBoxNo
    )
    
    if reply == MessageBoxNo:
        for item in items:
            colorize_item(item, None)
        tree_widget.clearSelection()
        return

    #-------------------------------------------------------------
    interpolationDict = itemInterpolation.data(0, Qt.ItemDataRole.UserRole)
    X1Coords = interpolationDict['X1Coords']
    X2Coords = interpolationDict['X2Coords']
    f_1to2, f_2to1 = defineInterpolationWindow.defineInterpolationFunctions(X1Coords, X2Coords, interpolationMode=interpolationMode)

    for item in itemSeries_selected:
        seriesDict = item.data(0, Qt.ItemDataRole.UserRole)
        series = seriesDict['Series']
        
        #series = series.groupby(series.index).mean()

        # to keep replicates
        series_interpolated = pd.Series(series.values, index=f_2to1(series.index))

        interpolated_Id = generate_Id()
        interpolated_seriesDict = seriesDict | {'Id': interpolated_Id,
            'Type': 'Series interpolated',
            'Series': series_interpolated,
            'InterpolationMode': interpolationMode,
            'X': interpolationDict['X1Name'], 
            'XOriginal': seriesDict['X'], 
            'XOriginalValues': series.index.to_list(),
            'X1Coords': X1Coords,
            'X2Coords': X2Coords, 
            'Color': generate_color(exclude_color=seriesDict['Color']),
            'Date': datetime.datetime.now().strftime("Created %Y/%m/%d at %H:%M:%S"),
            'History': append_to_htmlText(seriesDict['History'], 
                f'Series <i><b>{seriesDict["Id"]}</i></b> interpolated with INTERPOLATION <i><b>{interpolationDict["Id"]}</i></b> with mode {interpolationMode}<BR>---> series <i><b>{interpolated_Id}</b></i>'),
            'Comment': ''
        }

        ws_item = item.parent()
        position = ws_item.indexOfChild(item)
        add_item_tree_widget(ws_item, interpolated_seriesDict, position+1)

    #-------------------------------------------------------------
    main_window.setFocus()                  # replace selection
    tree_widget.clearSelection()
    for item in itemSeries_selected + itemInterpolations_selected:
        colorize_item(item, None)
        item.setSelected(True)

#========================================================================================
def close_all_windows():
    global open_displayWindows

    for Id_displayWindow in list(open_displayWindows.keys()):
        open_displayWindows[Id_displayWindow].close()
    open_displayWindows.clear()

#========================================================================================
def delete_parent_node(item):
    global open_ws

    index = tree_widget.indexOfTopLevelItem(item)
    tree_widget.takeTopLevelItem(index)
    del open_ws[id(item)]

#========================================================================================
def move_WorkSheet(direction):
    selected_item = tree_widget.currentItem()

    if selected_item and not selected_item.parent():  # Ensure it's a top-level parent
        index = tree_widget.indexOfTopLevelItem(selected_item)
        new_index = index + direction

        if 0 <= new_index < tree_widget.topLevelItemCount():  # Check if the new position is valid
    
            tree_widget.blockSignals(True)

            # Preserve the expanded/collapsed state
            was_expanded = selected_item.isExpanded()

            # Extract the selected item
            selected_item = tree_widget.takeTopLevelItem(index)

            # Store children data before removing them
            children_data = []
            for i in range(selected_item.childCount()):
                child = selected_item.child(i)
                itemDict = child.data(0, Qt.ItemDataRole.UserRole)  # Get the data for the child
                children_data.append(itemDict)

            # Remove all children to ensure a fresh start
            selected_item.takeChildren()

            # Insert the parent at the new position
            tree_widget.insertTopLevelItem(new_index, selected_item)

            # Restore children using add_item_tree_widget
            for itemDict in children_data:
                add_item_tree_widget(selected_item, itemDict, mark=False)       # do not mark as to be saved

            # Restore expanded/collapsed state
            selected_item.setExpanded(was_expanded)

            # Ensure it's still selected after moving
            tree_widget.setCurrentItem(selected_item)

            tree_widget.blockSignals(False)

#========================================================================================
def show_context_menu(point):
    item = tree_widget.itemAt(point)
    if item and not item.parent():  # Only allow delete on parents (ws)
        context_menu = QMenu(tree_widget)
        up_action = context_menu.addAction("Up")
        down_action = context_menu.addAction("Down")
        context_menu.addSeparator()
        delete_action = context_menu.addAction("Remove")
        action = context_menu.exec(tree_widget.mapToGlobal(point))
        if action == delete_action:
            delete_parent_node(item)
        elif action == up_action:
            move_WorkSheet(-1)
        elif action == down_action:
            move_WorkSheet(1)

#========================================================================================
def count_unsaved_ws():
    count = 0
    for ws_item in tree_widget.get_parents():
        if ws_item.data(0, Qt.ItemDataRole.UserRole):
            count += 1
    return count

#========================================================================================
def is_item_in_ws(ws_item, child_item):
    child_itemDict = child_item.data(0, Qt.ItemDataRole.UserRole)
    for i in range(ws_item.childCount()):
        item = ws_item.child(i)
        itemDict = item.data(0, Qt.ItemDataRole.UserRole)
        if child_itemDict['Id'] == itemDict['Id']:
            return True
    return False

#========================================================================================
def mark_ws(ws_item):
    tree_widget.blockSignals(True)
    ws_item.setData(0, Qt.ItemDataRole.UserRole, True)  # Mark ws_item as modified
    if not ws_item.text(0).endswith(' *'):
        ws_item.setText(0, f"{ws_item.text(0)} *") 
    tree_widget.blockSignals(False)

#========================================================================================
def unmark_ws(ws_item):
    tree_widget.blockSignals(True)
    ws_item.setData(0, Qt.ItemDataRole.UserRole, False)  # Reset modification state
    ws_item.setText(0, ws_item.text(0).replace(" *", ""))  # Remove visual cue
    tree_widget.blockSignals(False)
    
#========================================================================================
def remark_ws(ws_item):
    tree_widget.blockSignals(True)
    if ws_item.data(0, Qt.ItemDataRole.UserRole) and not ws_item.text(0).endswith(' *'):
        ws_item.setText(0, f"{ws_item.text(0)} *") 
    tree_widget.blockSignals(False)

#========================================================================================
def copy_items():
    selected_items = tree_widget.selectedItems()
    tree_widget.clipboard_items = selected_items

#========================================================================================
def cut_items():
    selected_items = tree_widget.selectedItems()
    tree_widget.clipboard_items = selected_items

    for item in selected_items:
        ws_item = item.parent()
        ws_item.removeChild(item)
        mark_ws(ws_item)

#========================================================================================
def paste_items():
    target_item = tree_widget.currentItem()

    if not target_item:             # no ws present
        return

    ws_item = target_item.parent() if target_item.parent() else target_item
    position = ws_item.indexOfChild(target_item)

    for item in tree_widget.clipboard_items:
        if is_item_in_ws(ws_item, item):
            main_window.statusBar().showMessage('Item(s) already in', 5000)
            continue
        itemDict = item.data(0, Qt.ItemDataRole.UserRole)
        add_item_tree_widget(ws_item, itemDict, position+1, update=False)

#========================================================================================
def on_item_double_clicked(item, column):

    tree_widget.blockSignals(True)

    if not item.parent(): item_isWS = True
    else: item_isWS = False

    itemDict = item.data(0, Qt.ItemDataRole.UserRole)

    if column == 0:
        item.setFlags(item.flags() | ItemIsEditable)
        if item_isWS:
            item.setText(0, item.text(0).replace(" *", ""))  # Remove visual cue
    elif column == 2:
        tree_widget.custom_tooltip.hide()
        item.setFlags(item.flags() | ItemIsEditable)
    elif column == 3 and (itemDict['Type'].startswith('Series') or
                          itemDict['Type'] == "INTERPOLATION"): 
        item.setFlags(item.flags() | ItemIsEditable)
    elif column == 4 and itemDict['Type'].startswith('Series'):
        item.setFlags(item.flags() | ItemIsEditable)
    else:
        item.setFlags(item.flags() & ~ItemIsEditable)

    tree_widget.blockSignals(False)

#========================================================================================
def show_dialog(title, fileHTML, width, height):
    with open(fileHTML, 'r', encoding='utf-8') as file:
        html_text = file.read()
    
    dialog = QDialog()
    dialog.setWindowTitle(title)
    dialog.setFixedSize(width, height)
    
    main_layout = QVBoxLayout()
    dialog.setLayout(main_layout)
    text_browser = QTextBrowser()
    text_browser.setHtml(html_text)
    text_browser.setOpenExternalLinks(True)
    main_layout.addWidget(text_browser)
    
    button_layout = QHBoxLayout()
    button_layout.addItem(QSpacerItem(40, 20, QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Minimum))
    ok_button = QPushButton('OK')
    ok_button.clicked.connect(dialog.accept)
    icon = QApplication.style().standardIcon(QStyle.StandardPixmap.SP_DialogApplyButton)
    ok_button.setIcon(icon)
    button_layout.addWidget(ok_button)
    main_layout.addLayout(button_layout)

    dialog.exec()

#========================================================================================
def exit_confirm(parent):

    if count_unsaved_ws() == 0:
        reply = QMessageBox.question(
            parent, "Exit",
            "Are you sure you want to exit the application?",
            MessageBoxYes | MessageBoxNo,
            MessageBoxNo
        )
        return reply == MessageBoxYes

    else:
        msg = QMessageBox(parent)
        msg.setWindowTitle("Exit")
        msg.setText("There are unsaved worksheets. What do you want to do?")
        no_save_exit = msg.addButton(MessageBoxDiscard)
        cancel = msg.addButton(MessageBoxCancel)
        save_exit = msg.addButton("Save all unsaved worksheets and Exit", MessageBoxAcceptRole)
        msg.exec()

        clicked = msg.clickedButton()
        if clicked == save_exit:
            #print("save_exit")
            save_WorkSheets()
            return True
        elif clicked == no_save_exit:
            #print("no_save_exit")
            return True 
        else:
            #print("cancel")
            return False 

#========================================================================================
class MainWindow(QMainWindow):
    def closeEvent(self, event):
        if exit_confirm(self):
            event.accept()
            QApplication.instance().quit()
        else:
            event.ignore()

#========================================================================================
def open_preferencesDialog():
    dlg = preferencesDialog(on_font_changed, on_custom_tooltip_changed, main_window)
    dlg.show()

#========================================================================================
def on_font_changed(size):
    font = app.font()
    font.setPointSize(size)
    app.setFont(font)

    update_tree_visuals(tree_widget)

    settings = QSettings("MyPythonApps", "PyAnalySeries")
    settings.setValue("ui/fontSize", size)

#========================================================================================
def on_custom_tooltip_changed(enabled):
    tree_widget.set_custom_tooltip_enabled(enabled)

    settings = QSettings("MyPythonApps", "PyAnalySeries")
    settings.setValue("ui/customTooltipEnabled", enabled)

#========================================================================================
app = QApplication(sys.argv)
app.setStyle("Fusion")

app_dir = Path(__file__).resolve().parent

settings = QSettings("MyPythonApps", "PyAnalySeries")
fontSize = settings.value("ui/fontSize", 12, type=int)
enabled = settings.value("ui/customTooltipEnabled", True, type=bool)

fontArial = QFont('Arial', fontSize)
app.setFont(fontArial)

icon = QIcon(str(app_dir / 'resources' / 'PyAnalySeries_icon.svg'))
app.setWindowIcon(icon)

main_window = MainWindow()
main_window.setWindowTitle(f"PyAnalySeries {version}")
main_window.setGeometry(100, 100, 1400, 600)

main_widget = QWidget()
layout = QVBoxLayout()

tree_widget = create_tree_widget()
tree_widget.setContextMenuPolicy(CustomContextMenu)
tree_widget.customContextMenuRequested.connect(show_context_menu)
tree_widget.itemChanged.connect(on_item_changed)
tree_widget.itemDoubleClicked.connect(on_item_double_clicked)
tree_widget.set_custom_tooltip_enabled(enabled)
update_tree_visuals(tree_widget)

layout.addWidget(tree_widget)

main_widget.setLayout(layout)
main_window.setCentralWidget(main_widget)

menu_bar = main_window.menuBar()

#----------------------------------------------
file_menu = menu_bar.addMenu("File")

newWS_action = QAction("New worksheet", main_window)
newWS_action.setShortcut('Ctrl+N')
newWS_action.triggered.connect(new_WorkSheet)
openWS_action = QAction("Open worksheet(s)", main_window)
openWS_action.setShortcut('Ctrl+O')
openWS_action.triggered.connect(open_WorkSheet)
saveWS_action = QAction("Save current worksheet", main_window)
saveWS_action.setShortcut('Ctrl+S')
saveWS_action.triggered.connect(save_WorkSheetCurrent)
saveWSs_action = QAction("Save all worksheets", main_window)
saveWSs_action.setShortcut('Ctrl+Shift+S')
saveWSs_action.triggered.connect(save_WorkSheets)
exit_action = QAction('Exit', main_window)
exit_action.setShortcut(QKeySequence.StandardKey.Quit)
exit_action.triggered.connect(main_window.close)

file_menu.addAction(newWS_action)
file_menu.addAction(openWS_action)
file_menu.addAction(saveWS_action)
file_menu.addAction(saveWSs_action)
file_menu.addSeparator()
file_menu.addAction(exit_action)

#----------------------------------------------
edit_menu = menu_bar.addMenu("Edit")

cut_action = QAction("Cut", main_window)
#cut_action.setShortcuts([QKeySequence("Ctrl+x"), QKeySequence(Qt.Key.Key_Delete)])
cut_action.setShortcut(QKeySequence("Ctrl+x"))
cut_action.triggered.connect(cut_items)

copy_action = QAction("Copy", main_window)
copy_action.setShortcut(QKeySequence("Ctrl+c"))
copy_action.triggered.connect(copy_items)

paste_action = QAction("Paste", main_window)
paste_action.setShortcut(QKeySequence("Ctrl+v"))
paste_action.triggered.connect(paste_items)

edit_menu.addAction(cut_action)
edit_menu.addAction(copy_action)
edit_menu.addAction(paste_action)

#----------------------------------------------
create_menu = menu_bar.addMenu("Create")

importData_action = QAction("Import data", main_window)
importData_action.setShortcut('Ctrl+m')
importData_action.triggered.connect(import_Data)

randomSeries_action = QAction("Random series", main_window)
randomSeries_action.triggered.connect(define_randomSeries)
insolationAstroSeries_action = QAction("Insolation / Astronomical series", main_window)
insolationAstroSeries_action.triggered.connect(define_insolationAstroSeries)
sinusoidalSeries_action = QAction("Sinusoidal series", main_window)
sinusoidalSeries_action.triggered.connect(define_sinusoidalSeries)

create_menu.addAction(importData_action)
create_menu.addSeparator()
create_menu.addAction(randomSeries_action)
create_menu.addAction(insolationAstroSeries_action)
create_menu.addAction(sinusoidalSeries_action)

#----------------------------------------------
display_menu = menu_bar.addMenu("Display")

display_action = QAction("Display Single", main_window)
display_action.setShortcut('Ctrl+d')
display_action.triggered.connect(displaySingleSeries_selected_series)

displayTogetherSeries_action = QAction("Display Together", main_window)
displayTogetherSeries_action.setShortcut('Ctrl+t')
displayTogetherSeries_action.triggered.connect(lambda: displayMultipleSeries_selected_series(overlaid=True))

displayStackedSeries_action = QAction("Display Stacked", main_window)
displayStackedSeries_action.setShortcut('Ctrl+k')
displayStackedSeries_action.triggered.connect(lambda: displayMultipleSeries_selected_series(overlaid=False))

close_all_action = QAction("Close all Display windows")
close_all_action.triggered.connect(close_all_windows)

display_menu.addAction(display_action)
display_menu.addAction(displayTogetherSeries_action)
display_menu.addAction(displayStackedSeries_action)
display_menu.addAction(close_all_action)

def add_section(menu, text, first=False):
    container = QWidget()
    layout = QVBoxLayout(container)
    layout.setContentsMargins(12, 6, 12, 6)
    layout.setSpacing(4)

    # --- Top line (only if not first section) ---
    if not first:
        top_line = QFrame()
        top_line.setFixedHeight(2)
        top_line.setStyleSheet("background-color: #999;")
        layout.addWidget(top_line)

    # --- Label ---
    label = QLabel(text)
    label.setFont(menu.font())
    layout.addWidget(label)

    action = QWidgetAction(menu)
    action.setDefaultWidget(container)
    menu.addAction(action)

#----------------------------------------------
process_menu = menu_bar.addMenu("Process")

# ----------------
add_section(process_menu, "Transforms", first=True)

computeAggregate_action = QAction("Edit / Aggregate / Clean ...", main_window)
computeAggregate_action.setShortcut("Ctrl+e")
computeAggregate_action.triggered.connect(compute_aggregate)
process_menu.addAction(computeAggregate_action)

process_menu.addSeparator()

computeDetrend_action = QAction("Detrend ...", main_window)
computeDetrend_action.triggered.connect(compute_detrend)
process_menu.addAction(computeDetrend_action)

process_menu.addSeparator()

defineSampling_action = QAction("Define Sampling ...", main_window)
defineSampling_action.setShortcut("Ctrl+a")
defineSampling_action.triggered.connect(define_sampling)
process_menu.addAction(defineSampling_action)

applySampling_action = QAction("Apply Sampling", main_window)
applySampling_action.triggered.connect(apply_sampling)
process_menu.addAction(applySampling_action)

# ----------------
add_section(process_menu, "Interpolation")

defineInterpolation_action = QAction("Define Interpolation ...", main_window)
defineInterpolation_action.setShortcut("Ctrl+i")
defineInterpolation_action.triggered.connect(define_interpolation)
process_menu.addAction(defineInterpolation_action)

applyInterpolationLinear_action = QAction("Apply Interpolation Linear", main_window)
applyInterpolationLinear_action.triggered.connect(lambda: apply_interpolation("Linear"))
process_menu.addAction(applyInterpolationLinear_action)

applyInterpolationPCHIP_action = QAction("Apply Interpolation PCHIP", main_window)
applyInterpolationPCHIP_action.setToolTip("This action applies PCHIP interpolation to the selected data.")
applyInterpolationPCHIP_action.triggered.connect(lambda: apply_interpolation("PCHIP"))
process_menu.addAction(applyInterpolationPCHIP_action)

# ----------------
add_section(process_menu, "X-Domain Filter")

defineFilter_action = QAction("Define Smoothing Average ...", main_window)
defineFilter_action.setShortcut("Ctrl+f")
defineFilter_action.triggered.connect(define_filter)
process_menu.addAction(defineFilter_action)

applyFilter_action = QAction("Apply Smoothing Average", main_window)
applyFilter_action.triggered.connect(apply_filter)
process_menu.addAction(applyFilter_action)

# ----------------
add_section(process_menu, "Frequency-Domain Filter")

computeFreqFilter_action = QAction("Frequency Filter ...", main_window)
computeFreqFilter_action.triggered.connect(compute_frequencyFilter)
process_menu.addAction(computeFreqFilter_action)

# ----------------
add_section(process_menu, "Spectral Estimation")

computePowerSpectrum_action = QAction("Power Spectrum (PSD) ...", main_window)
computePowerSpectrum_action.triggered.connect(compute_powerSpectrum)
process_menu.addAction(computePowerSpectrum_action)

#----------------------------------------------
settings_menu = menu_bar.addMenu('Settings')

preferences_action = QAction("Preferences", main_window)
preferences_action.triggered.connect(open_preferencesDialog)

settings_menu.addAction(preferences_action)

#----------------------------------------------
help_menu = menu_bar.addMenu('Help')

help_action = QAction('Help', main_window)
help_action.triggered.connect(lambda: show_dialog('Help', app_dir / 'resources' / 'help.html', 1000, 800))
help_menu.addAction(help_action)

#----------------------------------------------
about_menu = menu_bar.addMenu('About')

about_action = QAction('About', main_window)
about_action.triggered.connect(lambda: show_dialog('About', app_dir / 'resources' / 'about.html', 1000, 600))
about_menu.addAction(about_action)

#----------------------------------------------
if filesName:
    for fileName in filesName: 
        print('Loading...', fileName)
        load_WorkSheet(fileName)

#----------------------------------------------
main_window.setStatusBar(QStatusBar())
main_window.statusBar().showMessage('Application ready', 5000)
main_window.show()

sys.exit(app_exec(app))
